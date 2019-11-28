import openpyxl
# from cpm.models import GoodsInfo


def open_excel(path):
    wb = openpyxl.load_workbook(path)
    sheetnames = wb.sheetnames  # ['sheet1','sheet2']
    sheet = wb[sheetnames[0]] # <Worksheet "sheet1">
    rows = sheet.max_row  # 获取sheet的最大行数
    cols = sheet.max_column  # 获取sheet的最大列数

    # 通过名字访问Cell对象, 通过value属性获取值
    # 通过行和列确定数据
    # a12 = sheet.cell(row=1, column=2).value

    # 获取一列数据, sheet.iter_rows() 获取所有的行
    # """
    # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
    # """
    # for one_column_data in sheet.iter_rows():
    #     print(one_column_data[0].value)

    # # 获取一行数据, sheet.iter_cols() 获取所有的列
    # """
    # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
    # """
    # for row in range(0,rows):
    #     for one_row_data in sheet.iter_cols(row):
    #         print(one_row_data[row].value, end="\t")

    arr = []
    for i,cells in enumerate(sheet.rows):
        temp_dict = {}
        # if i >10:
        #     return
        # print(temp) # tuple (Cell,Cell,Cell...)
        if cells[1].value != None:

            temp_dict['name'] = cells[1].value
            temp_dict['weight'] = cells[7].value
            arr.append(temp_dict)
    print(arr)
        
 
def main():
    open_excel('/root/下载/商品信息缩小版.xlsx')

if __name__ == "__main__":
    main()